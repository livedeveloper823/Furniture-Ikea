from time import sleep
import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
import openpyxl, json, re


def wait_url(driver: webdriver.Chrome, url: str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)

def find_element(driver: webdriver.Chrome, whichBy, unique: str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique: str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return elements

wb = openpyxl.Workbook()

ws = wb.active

ws.merge_cells('G1:AO1')
ws.merge_cells('AP1:BH1')

ws['G1'] = 'MATERIALES Y CARACTERÍSTICAS'
ws['AP1'] = 'Medidas'

ws['A1'] = "Categoría"
ws['B1'] = "Línea"
ws['C1'] = "Sublinea"
ws['D2'] = "Precio"
ws['E2'] = "Nombre"
ws['F2'] = "Atributos3"
ws['G2'] = "Atributos4"
ws['H2'] = "Atributos5"
ws['I2'] = "Atributos6"
ws['J2'] = "Atributos7"
ws['K2'] = "Atributos8"
ws['L2'] = "Atributos9"
ws['M2'] = "Atributos10"
ws['N2'] = "Atributos11"
ws['O2'] = "Atributos11"
ws['P2'] = "Atributos11"
ws['Q2'] = "Atributos11"
ws['R2'] = "Atributos11"
ws['S2'] = "Atributos11"
ws['T2'] = "Atributos11"
ws['U2'] = "Atributos11"
ws['V2'] = "Atributos11"
ws['W2'] = "Atributos1"
ws['X2'] = "Atributos2"
ws['Y2'] = "Atributos3"
ws['Z2'] = "Atributos4"
ws['AA2'] = "Atributos5"
ws['AB2'] = "Atributos6"
ws['AC2'] = "Atributos7"
ws['AD2'] = "Atributos8"
ws['AE2'] = "Atributos9"
ws['AF2'] = "Atributos10"
ws['AG2'] = "Atributos11"
ws['AH2'] = "Atributos11"
ws['AI2'] = "Atributos11"
ws['AJ2'] = "Atributos11"
ws['AK2'] = "Atributos11"
ws['AL2'] = "Atributos11"
ws['AM2'] = "Atributos11"
ws['AN2'] = "Atributos11"
ws['AO2'] = "Atributos11"
ws['AP2'] = "Atributos5"
ws['AQ2'] = "Atributos6"
ws['AR2'] = "Atributos7"
ws['AS2'] = "Atributos8"
ws['AT2'] = "Atributos9"
ws['AU2'] = "Atributos10"
ws['AV2'] = "Atributos11"
ws['AW2'] = "Atributos11"
ws['AX2'] = "Atributos11"
ws['AY2'] = "Atributos11"
ws['AZ2'] = "Atributos11"
ws['BA2'] = "Atributos11"
ws['BB2'] = "Atributos11"
ws['BC2'] = "Atributos11"
ws['BD2'] = "Atributos5"
ws['BE2'] = "Atributos6"
ws['BF2'] = "Atributos7"
ws['BG2'] = "Atributos8"
ws['BH2'] = "Atributos8"


wb.save(f'ikea.xlsx')

urls = [
  "https://www.ikea.com/co/es/cat/sofas-tela-10661/",
  "https://www.ikea.com/co/es/cat/sofas-cuero-ecocuero-10662/",
  "https://www.ikea.com/co/es/cat/sofas-modulares-16238/",
  "https://www.ikea.com/co/es/cat/sofas-cama-futones-10663/", 
  "https://www.ikea.com/co/es/cat/puffs-descansapies-20926/",
  "https://www.ikea.com/co/es/cat/mesas-comedor-21825/",
  "https://www.ikea.com/co/es/cat/sillas-comedor-25219/",
  "https://www.ikea.com/co/es/cat/juegos-comedor-19145/",
  "https://www.ikea.com/co/es/cat/butacos-22659/",
  "https://www.ikea.com/co/es/cat/sillas-barra-20864/",
  "https://www.ikea.com/co/es/cat/camas-dobles-16284/",
  "https://www.ikea.com/co/es/cat/camas-almacenaje-25205/",
  "https://www.ikea.com/co/es/cat/camas-sencillas-16285/",
  "https://www.ikea.com/co/es/cat/cunas-bebes-45781/",
  "https://www.ikea.com/co/es/cat/cabeceros-cama-19064/"
]
match_num = 0

driver = webdriver.Chrome()
driver.maximize_window()

for url in urls:
    driver.get(url)
    wait_url(driver, url)

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2)")
    sleep(0.3)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight*3/4)")
    sleep(5)

    while True:
        try:
            driver.find_element(By.CLASS_NAME, "plp-btn--secondary").click()
            sleep(3)
        except:
            break
        sleep(0.1)

    products = find_elements(driver, By.CLASS_NAME, "plp-fragment-wrapper")
    product_urls = []
    
    for product in products:
        product_url = product.find_element(By.TAG_NAME, "a").get_attribute("href")
        product_urls.append(product_url)
        print(product_url)

    for product_url in product_urls:
        match_num += 1
        workbook = openpyxl.load_workbook("ikea.xlsx")
        sheet = workbook['Sheet']
        driver.get(product_url)
        sleep(0.5)
        product_name = find_element(driver, By.CLASS_NAME, "js-price-package").find_element(By.CLASS_NAME, "pip-header-section__title--big").text
        print(product_name)
        sheet[f'E{match_num+1}'] = product_name
        product_subncription = find_element(driver, By.CLASS_NAME, "js-price-package").find_element(By.CLASS_NAME, "pip-header-section__description").text
        product_price = find_element(driver, By.CLASS_NAME, "js-price-package").find_element(By.CLASS_NAME, "pip-temp-price--currency-super-aligned").find_element(By.CLASS_NAME, "pip-temp-price__integer").text.replace(".", "")
        print(product_price)
        sheet[f'D{match_num+1}'] = product_price
        find_element(driver, By.CLASS_NAME, "pip-product__left-bottom").find_element(By.CLASS_NAME, "js-product-information-section").find_element(By.ID, "pip-product-information-section-list-0").click()
        sleep(2)
        find_element(driver, By.ID, "product-details-material-and-care").click()
        sleep(0.5)
        material_data_box = find_element(driver, By.ID, "SEC_product-details-material-and-care")
        material_data = material_data_box.find_elements(By.TAG_NAME, "dl")

        # Structure
        try:
            structure = material_data_box.find_element(By.CLASS_NAME, "pip-product-details__container").find_element(By.CLASS_NAME, "pip-product-details__material-sub-header").text
            print(structure)
            sheet[f'F{match_num+1}'] = structure
        except:
            pass

        # Materials
        for each_material in material_data:
            try:
                each_material_text = each_material.find_element(By.TAG_NAME, "dt").text
                each_material_text1 = each_material.find_element(By.TAG_NAME, "dd").text
                if each_material_text == "Tela:":
                    material1 = each_material_text1
                    sheet[f'G{match_num+1}'] = material1
                    print(material1)
                elif each_material_text == "Estructura del respaldo y del asiento:":
                    material2 = each_material_text1
                    sheet[f'H{match_num+1}'] = material2
                    print(material2)
                elif each_material_text == "Estructura de reposabrazos:":
                    material3 = each_material_text1
                    sheet[f'I{match_num+1}'] = material3
                    print(material3)
                elif each_material_text == "Cojín de respaldo:":
                    material4 = each_material_text1
                    sheet[f'J{match_num+1}'] = material4
                    print(material4)
                elif each_material_text == "Cojín de asiento:":
                    material5 = each_material_text1
                    sheet[f'K{match_num+1}'] = material5
                    print(material5)
                elif each_material_text == "Pata:":
                    material6 = each_material_text1
                    sheet[f'L{match_num+1}'] = material6
                    print(material6)
                elif each_material_text == "Cuero:":
                    material7 = each_material_text1
                    sheet[f'M{match_num+1}'] = material7
                elif each_material_text == "Marco:":
                    material8 = each_material_text1
                    sheet[f'N{match_num+1}'] = material8
                elif each_material_text == "Forro:":
                    material9 = each_material_text1
                    sheet[f'O{match_num+1}'] = material9
                elif each_material_text == "Respaldo:":
                    material10 = each_material_text1
                    sheet[f'P{match_num+1}'] = material10
                elif each_material_text == "Tejido posterior:":
                    material11 = each_material_text1
                    sheet[f'Q{match_num+1}'] = material11
                    print(material11)
                elif each_material_text == "Cojín de asiento:":
                    material12 = each_material_text1
                    sheet[f'R{match_num+1}'] = material12
                elif each_material_text == "Compás para puerta:":
                    material13 = each_material_text1
                    sheet[f'S{match_num+1}'] = material13
                elif each_material_text == "Clip:":
                    material14 = each_material_text1
                    sheet[f'T{match_num+1}'] = material14
                elif each_material_text == "Travesaño:":
                    material15 = each_material_text1
                    sheet[f'U{match_num+1}'] = material15
                elif each_material_text == "Estructura de respaldo:":
                    material16 = each_material_text1
                    sheet[f'V{match_num+1}'] = material16
                elif each_material_text == "Estructura de asiento:":
                    material17 = each_material_text1
                    sheet[f'W{match_num+1}'] = material17
                elif each_material_text == "Estructura de reposabrazos:":
                    material18 = each_material_text1
                    sheet[f'X{match_num+1}'] = material18
                elif each_material_text == "Cajón para cama:":
                    material19 = each_material_text1
                    sheet[f'Y{match_num+1}'] = material19
                elif each_material_text == "Pata/ Borde largo:":
                    material20 = each_material_text1
                    sheet[f'Z{match_num+1}'] = material20
                elif each_material_text == "Pata interior:":
                    material21 = each_material_text1
                    sheet[f'AA{match_num+1}'] = material21
                elif each_material_text == "Tabla:":
                    material22 = each_material_text1
                    sheet[f'AB{match_num+1}'] = material22
                elif each_material_text == "Tejido de base:":
                    material23 = each_material_text1
                    sheet[f'AC{match_num+1}'] = material23
                elif each_material_text == "Asiento:":
                    material24 = each_material_text1
                    sheet[f'AD{match_num+1}'] = material24
                elif each_material_text == "Asiento/ Respaldo:":
                    material25 = each_material_text1
                    sheet[f'AE{match_num+1}'] = material25
                elif each_material_text == "Estructura de pata/ Riel transversal:":
                    material26 = each_material_text1
                    sheet[f'AF{match_num+1}'] = material26
                elif each_material_text == "Asiento/ Dorso:":
                    material27 = each_material_text1
                    sheet[f'AG{match_num+1}'] = material27
                elif each_material_text == "Cabecera/pies de cama:":
                    material28 = each_material_text1
                    sheet[f'AH{match_num+1}'] = material28
                elif each_material_text == "Lateral de cama:":
                    material29 = each_material_text1
                    sheet[f'AI{match_num+1}'] = material29
                elif each_material_text == "Tablillas de cama:":
                    material30 = each_material_text1
                    sheet[f'AJ{match_num+1}'] = material30
                elif each_material_text == "Cinta:":
                    material31 = each_material_text1
                    sheet[f'AK{match_num+1}'] = material31
                elif each_material_text == "Componentes principales/ Pared divisora/ Frente de cajón:":
                    material32 = each_material_text1
                    sheet[f'AL{match_num+1}'] = material32
                elif each_material_text == "Fondo de cajón:":
                    material33 = each_material_text1
                    sheet[f'AM{match_num+1}'] = material33
                elif each_material_text == "Relleno de la cabecera:":
                    material34 = each_material_text1
                    sheet[f'AN{match_num+1}'] = material34
                elif each_material_text == "Hilo:":
                    material35 = each_material_text1
                    sheet[f'AO{match_num+1}'] = material35
            except:
                pass
        
        # Close the material modal
        find_element(driver, By.CLASS_NAME, "pip-modal-header__close").click()
        sleep(1)

        # Measurement
        try:
            find_element(driver, By.CLASS_NAME, "pip-product__left-bottom").find_element(By.CLASS_NAME, "js-product-information-section").find_element(By.ID, "pip-product-information-section-list-2").click()
            sleep(1)
            dimentions = find_element(driver, By.CLASS_NAME, "pip-modal-body").find_element(By.CLASS_NAME, "pip-product-dimensions").find_element(By.CLASS_NAME, "pip-product-dimensions__dimensions-container").find_elements(By.TAG_NAME, "p")
            for dimention in dimentions:
                if "Longitud" in dimention.text.split(":")[0].strip():
                    length = dimention.text.split(":")[1].strip()
                    print(length)
                    sheet[f'AP{match_num+1}'] = length
                elif "Ancho" in dimention.text.split(":")[0].strip():
                    width = dimention.text.split(":")[1].strip()
                    print(width)
                    sheet[f'AQ{match_num+1}'] = width
                elif "Altura" in dimention.text.split(":")[0].strip(): 
                    height = dimention.text.split(":")[1].strip()
                    print(height)
                    sheet[f'AR{match_num+1}'] = height
                elif "Fondo" in dimention.text.split(":")[0].strip():
                    back = dimention.text.split(":")[1].strip()
                    print(back)
                    sheet[f'AS{match_num+1}'] = back
                elif "Altura incl. cojines del espaldar" in dimention.text.split(":")[0].strip():
                        backheight = dimention.text.split(":")[1].strip()
                        print(backheight)
                        sheet[f'AT{match_num+1}'] = backheight
                elif "Ancho de cama" in dimention.text.split(":")[0].strip():
                    bedwidth = dimention.text.split(":")[1].strip()
                    print(bedwidth)
                    sheet[f'AU{match_num+1}'] = bedwidth
                elif "Ancho del asiento" in dimention.text.split(":")[0].strip():
                    seatwidth = dimention.text.split(":")[1].strip()
                    print(seatwidth)
                    sheet[f'AV{match_num+1}'] = seatwidth
                elif "Altura del asiento" in dimention.text.split(":")[0].strip():
                    seatheight = dimention.text.split(":")[1].strip()
                    print(seatheight)
                    sheet[f'AW{match_num+1}'] = seatheight
                elif "Fondo del asiento" in dimention.text.split(":")[0].strip():
                    seatback = dimention.text.split(":")[1].strip()
                    print(seatback)
                    sheet[f'AX{match_num+1}'] = seatback
                elif "Altura piecero" in dimention.text.split(":")[0].strip():
                    footboardheight = dimention.text.split(":")[1].strip()
                    print(footboardheight)
                    sheet[f'AY{match_num+1}'] = footboardheight
                elif "Altura cabecera" in dimention.text.split(":")[0].strip():
                    headboardheight = dimention.text.split(":")[1].strip()
                    print(headboardheight)
                    sheet[f'AZ{match_num+1}'] = headboardheight
                elif "Largo del colchón" in dimention.text.split(":")[0].strip():
                    matresslength = dimention.text.split(":")[1].strip()
                    print(matresslength)
                    sheet[f'BA{match_num+1}'] = matresslength
                elif "Ancho del colchón" in dimention.text.split(":")[0].strip():
                    matresswidth = dimention.text.split(":")[1].strip()
                    print(matresswidth)
                    sheet[f'BB{match_num+1}'] = matresswidth
                elif "Ancho de cajón (int)" in dimention.text.split(":")[0].strip():
                    drawerwidth = dimention.text.split(":")[1].strip()
                    print(drawerwidth)
                    sheet[f'BC{match_num+1}'] = drawerwidth
                elif "Fondo de cajón (int)" in dimention.text.split(":")[0].strip():
                    drawerdepth = dimention.text.split(":")[1].strip()
                    print(drawerdepth)
                    sheet[f'BD{match_num+1}'] = drawerdepth
                elif "Largo de cama" in dimention.text.split(":")[0].strip():
                    bedlength = dimention.text.split(":")[1].strip()
                    print(bedlength)
                    sheet[f'BF{match_num+1}'] = bedlength
                elif "Máximo soportado" in dimention.text.split(":")[0].strip():
                    maxsupport = dimention.text.split(":")[1].strip()
                    print(maxsupport)
                    sheet[f'BG{match_num+1}'] = maxsupport
                elif "Altura libre bajo cuna" in dimention.text.split(":")[0].strip():
                    undercribheight = dimention.text.split(":")[1].strip()
                    print(undercribheight)
                    sheet[f'BH{match_num+1}'] = undercribheight
                elif "Grosor" in dimention.text.split(":")[0].strip():
                    thickness = dimention.text.split(":")[1].strip()
                    print(thickness)
                    sheet[f'BI{match_num+1}'] = thickness

            find_element(driver, By.CLASS_NAME, "pip-modal-header__close").click()
        except:
            find_element(driver, By.CLASS_NAME, "pip-product__left-bottom").find_element(By.CLASS_NAME, "js-product-information-section").find_element(By.ID, "pip-product-information-section-list-1").click()
            sleep(1)
            dimentions = find_element(driver, By.CLASS_NAME, "pip-modal-body").find_element(By.CLASS_NAME, "pip-product-dimensions").find_element(By.CLASS_NAME, "pip-product-dimensions__dimensions-container").find_elements(By.TAG_NAME, "p")
            for dimention in dimentions:
                if "Longitud" in dimention.text.split(":")[0].strip():
                    length = dimention.text.split(":")[1].strip()
                    print(length)
                    sheet[f'AP{match_num+1}'] = length
                elif "Ancho" in dimention.text.split(":")[0].strip():
                    width = dimention.text.split(":")[1].strip()
                    print(width)
                    sheet[f'AQ{match_num+1}'] = width
                elif "Altura" in dimention.text.split(":")[0].strip(): 
                    height = dimention.text.split(":")[1].strip()
                    print(height)
                    sheet[f'AR{match_num+1}'] = height
                elif "Fondo" in dimention.text.split(":")[0].strip():
                    back = dimention.text.split(":")[1].strip()
                    print(back)
                    sheet[f'AS{match_num+1}'] = back
                elif "Altura incl. cojines del espaldar" in dimention.text.split(":")[0].strip():
                        backheight = dimention.text.split(":")[1].strip()
                        print(backheight)
                        sheet[f'AT{match_num+1}'] = backheight
                elif "Ancho de cama" in dimention.text.split(":")[0].strip():
                    bedwidth = dimention.text.split(":")[1].strip()
                    print(bedwidth)
                    sheet[f'AU{match_num+1}'] = bedwidth
                elif "Ancho del asiento" in dimention.text.split(":")[0].strip():
                    seatwidth = dimention.text.split(":")[1].strip()
                    print(seatwidth)
                    sheet[f'AV{match_num+1}'] = seatwidth
                elif "Altura del asiento" in dimention.text.split(":")[0].strip():
                    seatheight = dimention.text.split(":")[1].strip()
                    print(seatheight)
                    sheet[f'AW{match_num+1}'] = seatheight
                elif "Fondo del asiento" in dimention.text.split(":")[0].strip():
                    seatback = dimention.text.split(":")[1].strip()
                    print(seatback)
                    sheet[f'AX{match_num+1}'] = seatback
                elif "Altura piecero" in dimention.text.split(":")[0].strip():
                    footboardheight = dimention.text.split(":")[1].strip()
                    print(footboardheight)
                    sheet[f'AY{match_num+1}'] = footboardheight
                elif "Altura cabecera" in dimention.text.split(":")[0].strip():
                    headboardheight = dimention.text.split(":")[1].strip()
                    print(headboardheight)
                    sheet[f'AZ{match_num+1}'] = headboardheight
                elif "Largo del colchón" in dimention.text.split(":")[0].strip():
                    matresslength = dimention.text.split(":")[1].strip()
                    print(matresslength)
                    sheet[f'BA{match_num+1}'] = matresslength
                elif "Ancho del colchón" in dimention.text.split(":")[0].strip():
                    matresswidth = dimention.text.split(":")[1].strip()
                    print(matresswidth)
                    sheet[f'BB{match_num+1}'] = matresswidth
                elif "Ancho de cajón (int)" in dimention.text.split(":")[0].strip():
                    drawerwidth = dimention.text.split(":")[1].strip()
                    print(drawerwidth)
                    sheet[f'BC{match_num+1}'] = drawerwidth
                elif "Fondo de cajón (int)" in dimention.text.split(":")[0].strip():
                    drawerdepth = dimention.text.split(":")[1].strip()
                    print(drawerdepth)
                    sheet[f'BD{match_num+1}'] = drawerdepth
                elif "Largo de cama" in dimention.text.split(":")[0].strip():
                    bedlength = dimention.text.split(":")[1].strip()
                    print(bedlength)
                    sheet[f'BF{match_num+1}'] = bedlength
                elif "Máximo soportado" in dimention.text.split(":")[0].strip():
                    maxsupport = dimention.text.split(":")[1].strip()
                    print(maxsupport)
                    sheet[f'BG{match_num+1}'] = maxsupport
                elif "Altura libre bajo cuna" in dimention.text.split(":")[0].strip():
                    undercribheight = dimention.text.split(":")[1].strip()
                    print(undercribheight)
                    sheet[f'BH{match_num+1}'] = undercribheight
                elif "Grosor" in dimention.text.split(":")[0].strip():
                    thickness = dimention.text.split(":")[1].strip()
                    print(thickness)
                    sheet[f'BI{match_num+1}'] = thickness
            
            find_element(driver, By.CLASS_NAME, "pip-modal-header__close").click()
        workbook.save('ikea.xlsx')
    match_num += 2